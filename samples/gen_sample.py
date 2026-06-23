"""Generate a synthetic .xlsx that mirrors the user's hard-to-read workbooks.

Sheets: raw vertical data, a formula sheet with a merged 3-column "売上" header,
a value-pasted sheet, a border-drawn pseudo-header sheet, and a PowerQuery load
target. A DataMashup (PowerQuery M) part is injected post-hoc because openpyxl
cannot write PowerQuery.
"""
from __future__ import annotations

import base64
import io
import re
import struct
import zipfile

import openpyxl
from openpyxl.styles import Border, Side

SECTION_M = """section Section1;

shared #"Q_元データ" = let
    Source = Csv.Document(File.Contents("sales.csv"))
in
    Source;

shared #"Q_売上" = let
    Source = #"Q_元データ",
    Added = Table.AddColumn(Source, "前月差", each [当月] - [前月])
in
    Added;
"""


def build_datamashup(section_m: str) -> bytes:
    """Build the pre-base64 DataMashup bytes: version + package-length + package zip.

    The package is a zip whose only entry is Formulas/Section1.m. Real Excel
    appends Permissions/Metadata sections after the package; our parser only
    reads the package, so a trailing-empty structure is sufficient here.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("Formulas/Section1.m", section_m)
        z.writestr("[Content_Types].xml", _PACKAGE_CONTENT_TYPES)
    package = buf.getvalue()
    header = struct.pack("<ii", 0, len(package))
    return header + package


def datamashup_base64(section_m: str) -> str:
    return base64.b64encode(build_datamashup(section_m)).decode("ascii")


_PACKAGE_CONTENT_TYPES = (
    '<?xml version="1.0" encoding="utf-8"?>'
    '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
    '<Default Extension="m" ContentType="application/x-ms-m"/>'
    '<Default Extension="xml" ContentType="text/xml"/>'
    "</Types>"
)


def _item_xml(section_m: str) -> str:
    b64 = datamashup_base64(section_m)
    return (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<DataMashup xmlns="http://schemas.microsoft.com/DataMashup">'
        f"{b64}</DataMashup>"
    )


def _build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # Sheet 1: 生データ (raw, vertical/long format, no formulas)
    ws = wb.active
    ws.title = "生データ"
    ws.append(["月", "店舗", "売上"])
    for month, shop, val in [
        ("1月", "A店", 100), ("1月", "B店", 120),
        ("2月", "A店", 110), ("2月", "B店", 130),
    ]:
        ws.append([month, shop, val])

    # Sheet 2: 売上集計 (formula, merged 3-col header, references 生データ)
    ws = wb.create_sheet("売上集計")
    ws["A1"] = "店舗"
    ws["B1"] = "売上"            # merged across B:D
    ws.merge_cells("B1:D1")
    ws["B2"] = "当月"
    ws["C2"] = "前月"
    ws["D2"] = "前月差"
    ws["A3"] = "A店"
    ws["B3"] = 110
    ws["C3"] = 100
    ws["D3"] = "=B3-C3"          # 前月差 = 当月 - 前月
    ws["A4"] = "B店"
    ws["B4"] = "=生データ!C3"    # cross-sheet reference
    ws["C4"] = 120
    ws["D4"] = "=B4-C4"

    # Sheet 3: 値貼り付け (pasted, no formulas, dense values)
    ws = wb.create_sheet("値貼り付け")
    ws.append(["店舗", "当月", "前月", "前月差"])
    ws.append(["A店", 110, 100, 10])
    ws.append(["B店", 120, 130, -10])

    # Sheet 4: 罫線表頭 (border-drawn pseudo header, no merge)
    ws = wb.create_sheet("罫線表頭")
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"] = "売上"
    for col in ("A1", "B1", "C1"):
        ws[col].border = box
    ws["A2"] = "当月"
    ws["B2"] = "前月"
    ws["C2"] = "前月差"
    ws.append([])  # row 3 spacer handled by explicit rows below
    ws["A3"] = 110
    ws["B3"] = 100
    ws["C3"] = 10

    # Sheet 5: PQ出力 (PowerQuery load target placeholder)
    ws = wb.create_sheet("PQ出力")
    ws.append(["店舗", "当月", "前月", "前月差"])
    ws.append(["A店", 110, 100, 10])

    return wb


def _inject_datamashup(path: str, section_m: str) -> None:
    """Append a customXml DataMashup part into the already-written xlsx zip."""
    with zipfile.ZipFile(path) as zin:
        items = zin.infolist()
        data = {i.filename: zin.read(i.filename) for i in items}

    item_name = "customXml/item1.xml"
    data[item_name] = _item_xml(section_m).encode("utf-8")

    # Register content type for the customXml part.
    ct = data["[Content_Types].xml"].decode("utf-8")
    override = '<Override PartName="/customXml/item1.xml" ContentType="application/xml"/>'
    if override not in ct:
        ct = ct.replace("</Types>", override + "</Types>")
        data["[Content_Types].xml"] = ct.encode("utf-8")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, payload in data.items():
            zout.writestr(name, payload)


def generate(path: str) -> None:
    wb = _build_workbook()
    wb.save(path)
    _inject_datamashup(path, SECTION_M)


if __name__ == "__main__":
    import sys

    target = sys.argv[1] if len(sys.argv) > 1 else "sample.xlsx"
    generate(target)
    print(f"wrote {target}")
