import openpyxl

from xlsx_flow.parser.preview import cell_display, sheet_preview, column_preview
from xlsx_flow.parser.sheets import recover_headers
from xlsx_flow.parser.workbook import analyze
from samples.gen_sample import generate


def test_cell_display_formula_keeps_equals(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    ws = openpyxl.load_workbook(out)["売上集計"]
    assert cell_display(ws["D3"]) == "=B3-C3"


def test_cell_display_integer_floats_are_clean(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    ws = openpyxl.load_workbook(out)["生データ"]
    # 売上 column value 100 should render without a trailing ".0"
    assert cell_display(ws["C2"]) == "100"


def test_cell_display_none_is_empty():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert cell_display(ws["Z99"]) == ""


def test_sheet_preview_caps_rows_and_cols():
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 30):
        ws.append([f"r{r}c{c}" for c in range(1, 20)])
    prev = sheet_preview(ws, max_rows=6, max_cols=8)
    assert len(prev["cells"]) == 6
    assert all(len(row) == 8 for row in prev["cells"])
    assert prev["truncated"] is True
    assert prev["dims"] == {"rows": 29, "cols": 19}


def test_column_preview_collects_samples_and_formula_example(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    ws = openpyxl.load_workbook(out)["売上集計"]
    cols = recover_headers(ws)
    # 前月差 column (D, index 4) is a formula column.
    dcol = next(c for c in cols if c.col_index == 4)
    prev = column_preview(ws, dcol.col_index, dcol.data_start)
    assert prev["formula_example"] == "=B3-C3"
    assert prev["range"].startswith("D")


def test_analyze_attaches_preview_to_sheet_and_column_nodes(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    model = analyze(str(out))
    sheet_node = next(n for n in model.nodes if n.id == "sheet:生データ")
    assert "preview" in sheet_node.attrs
    assert sheet_node.attrs["preview"]["cells"]
    col_nodes = [n for n in model.nodes if n.type == "column" and n.attrs.get("sheet") == "売上集計"]
    assert any("preview" in n.attrs for n in col_nodes)
