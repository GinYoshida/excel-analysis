from xlsx_flow.parser.powerquery import extract_powerquery, _split_queries
from samples.gen_sample import generate, SECTION_M


def test_split_queries_finds_both_names():
    pairs = _split_queries(SECTION_M)
    names = {n for n, _ in pairs}
    assert "Q_元データ" in names
    assert "Q_売上" in names


def test_extract_finds_source_and_dependency(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    queries, warnings = extract_powerquery(str(out))
    by_name = {q.name: q for q in queries}
    assert "Q_元データ" in by_name
    assert "Q_売上" in by_name
    # Q_元データ reads a Csv source
    assert any("Csv.Document" in s or "File.Contents" in s for s in by_name["Q_元データ"].sources)
    # Q_売上 references Q_元データ
    assert "Q_元データ" in by_name["Q_売上"].refs


def test_missing_datamashup_returns_warning(tmp_path):
    import openpyxl
    out = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    wb.save(out)
    queries, warnings = extract_powerquery(str(out))
    assert queries == []
    assert warnings  # at least one warning, no exception


def test_malformed_datamashup_warns_without_raising(tmp_path):
    import zipfile
    import openpyxl
    from xlsx_flow.parser.powerquery import extract_powerquery

    src = tmp_path / "bad.xlsx"
    openpyxl.Workbook().save(src)
    # read existing entries, then rewrite the zip adding a bogus DataMashup item
    with zipfile.ZipFile(src) as zin:
        data = {i.filename: zin.read(i.filename) for i in zin.infolist()}
    data["customXml/item1.xml"] = (
        b'<?xml version="1.0"?>'
        b'<DataMashup xmlns="http://schemas.microsoft.com/DataMashup">'
        b'!!!not-valid-base64!!!</DataMashup>'
    )
    with zipfile.ZipFile(src, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, payload in data.items():
            zout.writestr(name, payload)

    queries, warnings = extract_powerquery(str(src))
    assert queries == []
    assert warnings  # a warning was recorded, no exception escaped


def test_extract_finds_datamashup_under_xl_prefix(tmp_path):
    import zipfile
    import openpyxl
    from xlsx_flow.parser.powerquery import extract_powerquery
    from samples.gen_sample import _item_xml, SECTION_M

    src = tmp_path / "xlprefix.xlsx"
    openpyxl.Workbook().save(src)
    with zipfile.ZipFile(src) as zin:
        data = {i.filename: zin.read(i.filename) for i in zin.infolist()}
    # Real Excel stores DataMashup under xl/customXml/, not bare customXml/
    data["xl/customXml/item1.xml"] = _item_xml(SECTION_M).encode("utf-8")
    with zipfile.ZipFile(src, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, payload in data.items():
            zout.writestr(name, payload)

    queries, warnings = extract_powerquery(str(src))
    names = {q.name for q in queries}
    assert "Q_元データ" in names
    assert "Q_売上" in names


def test_extract_finds_wb_entity_and_uri(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    queries, _ = extract_powerquery(str(out))
    by_name = {q.name: q for q in queries}
    assert "T_値貼付" in by_name["Q_テーブル取込"].wb_entities
    assert "sales.csv" in by_name["Q_元データ"].uris


def test_decompose_steps_orders_and_links(tmp_path):
    from xlsx_flow.parser.powerquery import decompose_steps
    out = tmp_path / "s.xlsx"
    generate(str(out))
    queries, _ = extract_powerquery(str(out))
    body = {q.name: q.m_code for q in queries}["Q_売上"]
    steps = decompose_steps(body)
    names = [s.name for s in steps]
    assert names == ["Source", "Added"]
    added = next(s for s in steps if s.name == "Added")
    assert "Source" in added.refs


def test_decompose_steps_handles_non_let():
    from xlsx_flow.parser.powerquery import decompose_steps
    assert decompose_steps('Excel.CurrentWorkbook(){[Name="X"]}[Content]') == []


def test_decompose_steps_ignores_commas_in_brackets():
    from xlsx_flow.parser.powerquery import decompose_steps
    body = 'let\n  Source = Table.FromRows({{1,2},{3,4}}),\n  Out = Source\nin\n  Out'
    steps = decompose_steps(body)
    assert [s.name for s in steps] == ["Source", "Out"]
