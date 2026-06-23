import zipfile
import base64
import struct

import openpyxl

from samples.gen_sample import build_datamashup, generate


def test_build_datamashup_roundtrips_section_m():
    section = 'section Section1;\nshared Q = 1;'
    raw = build_datamashup(section)
    # header: version(int32 LE), package length(int32 LE), then package zip
    version, length = struct.unpack_from("<ii", raw, 0)
    assert version == 0
    pkg = raw[8:8 + length]
    with zipfile.ZipFile(__import__("io").BytesIO(pkg)) as z:
        assert z.read("Formulas/Section1.m").decode("utf-8") == section


def test_generate_creates_five_sheets_with_expected_names(tmp_path):
    out = tmp_path / "sample.xlsx"
    generate(str(out))
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"生データ", "売上集計", "値貼り付け", "罫線表頭", "PQ出力"}


def test_generate_injects_datamashup(tmp_path):
    out = tmp_path / "sample.xlsx"
    generate(str(out))
    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        item = [n for n in names if n.startswith("customXml/item") and n.endswith(".xml")]
        assert item, f"no customXml item part found in {names}"
        xml = z.read(item[0]).decode("utf-8")
        assert "DataMashup" in xml


def test_generate_uri_has_merged_sales_header(tmp_path):
    out = tmp_path / "sample.xlsx"
    generate(str(out))
    wb = openpyxl.load_workbook(out)
    ws = wb["売上集計"]
    merged = [str(r) for r in ws.merged_cells.ranges]
    # "売上" spans 3 columns on the header row
    assert any(":" in r for r in merged), f"expected a merged range, got {merged}"
