from xlsx_flow.model import Model, Node


def test_model_to_dict_includes_notes():
    m = Model(file="x.xlsx")
    m.notes.append({"sheet": "S", "addr": "A1", "text": "説明メモ", "kind": "text"})
    d = m.to_dict()
    assert d["notes"] == [{"sheet": "S", "addr": "A1", "text": "説明メモ", "kind": "text"}]


def test_model_notes_defaults_empty():
    assert Model(file="x.xlsx").to_dict()["notes"] == []


import openpyxl
from openpyxl.comments import Comment

from xlsx_flow.parser.notes import extract_notes


def test_extract_notes_picks_long_text_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "手順"
    ws["A1"] = "短い"  # 20文字未満: 拾わない
    ws["A2"] = "このシートは毎月の売上を店舗別に集計するための作業表です"  # 長文: 拾う
    ws["A3"] = 12345  # 数値: 拾わない
    notes = extract_notes(ws, min_len=20)
    picked = {(n["addr"], n["kind"]) for n in notes}
    assert ("A2", "text") in picked
    assert all(n["addr"] != "A1" for n in notes)
    assert all(n["addr"] != "A3" for n in notes)
    assert all(n["sheet"] == "手順" for n in notes)


def test_extract_notes_picks_cell_comment():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "集計"
    ws["B2"] = 10
    ws["B2"].comment = Comment("前月比の計算に注意すること", "作成者")
    notes = extract_notes(ws, min_len=20)
    comments = [n for n in notes if n["kind"] == "comment"]
    assert comments and comments[0]["addr"] == "B2"
    assert "前月比" in comments[0]["text"]


from xlsx_flow.parser.workbook import analyze
from samples.gen_sample import generate


def test_analyze_collects_notes(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    model = analyze(str(out))
    texts = " ".join(n["text"] for n in model.notes)
    assert "月次売上" in texts        # 長文セル E1
    assert "税抜" in texts            # セルコメント C1
    assert all({"sheet", "addr", "text", "kind"} <= set(n) for n in model.notes)


def test_analyze_notes_respects_budget(tmp_path, monkeypatch):
    import xlsx_flow.parser.workbook as wbmod
    out = tmp_path / "s.xlsx"
    generate(str(out))
    monkeypatch.setattr(wbmod, "NOTES_BUDGET", 1)
    model = wbmod.analyze(str(out))
    assert len(model.notes) <= 1
    assert any("メモ" in w or "notes" in w.lower() for w in model.warnings)
