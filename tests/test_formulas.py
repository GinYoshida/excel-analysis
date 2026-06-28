import openpyxl

from xlsx_flow.parser.formulas import extract_references, cell_to_column_index
from samples.gen_sample import generate


def test_cell_to_column_index():
    assert cell_to_column_index("A1") == 1
    assert cell_to_column_index("C3") == 3
    assert cell_to_column_index("AA10") == 27


def test_extracts_cross_sheet_reference(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    ws = openpyxl.load_workbook(out)["売上集計"]
    refs = extract_references(ws)
    targets = {(r["target_sheet"], r["target_ref"]) for r in refs}
    assert ("生データ", "C3") in targets


def test_same_sheet_formula_ignored(tmp_path):
    out = tmp_path / "s.xlsx"
    generate(str(out))
    ws = openpyxl.load_workbook(out)["売上集計"]
    refs = extract_references(ws)
    # =B3-C3 has no sheet qualifier, must not appear
    assert all(r["target_sheet"] for r in refs)


def test_cell_to_column_index_handles_absolute_ref():
    assert cell_to_column_index("$C$3") == 3


def test_cell_to_column_index_raises_on_malformed():
    import pytest
    with pytest.raises(ValueError):
        cell_to_column_index("not-a-ref")


def test_extract_cell_dependencies_same_sheet(tmp_path):
    from xlsx_flow.parser.formulas import extract_cell_dependencies
    out = tmp_path / "s.xlsx"
    generate(str(out))
    ws = openpyxl.load_workbook(out)["売上集計"]
    deps = {d["src_cell"]: d["precedents"] for d in extract_cell_dependencies(ws)}
    # D3 = B3 - C3 (same-sheet, unqualified)
    d3 = {(p["sheet"], p["ref"]) for p in deps["D3"]}
    assert (None, "B3") in d3 and (None, "C3") in d3


def test_extract_cell_dependencies_cross_sheet(tmp_path):
    from xlsx_flow.parser.formulas import extract_cell_dependencies
    out = tmp_path / "s.xlsx"
    generate(str(out))
    ws = openpyxl.load_workbook(out)["売上集計"]
    deps = {d["src_cell"]: d["precedents"] for d in extract_cell_dependencies(ws)}
    # B4 = 生データ!C3 (cross-sheet, qualified)
    b4 = {(p["sheet"], p["ref"]) for p in deps["B4"]}
    assert ("生データ", "C3") in b4
