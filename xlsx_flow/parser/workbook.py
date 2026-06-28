"""Orchestrate all parsers and assemble the intermediate Model."""
from __future__ import annotations

import openpyxl

from xlsx_flow.model import Model, Node, Edge
from xlsx_flow.parser.sheets import classify_sheet, recover_headers
from xlsx_flow.parser.formulas import extract_references, cell_to_column_index
from xlsx_flow.parser.powerquery import extract_powerquery, decompose_steps
from xlsx_flow.parser.preview import sheet_preview, column_preview


def _col_id(sheet: str, header_path: list[str]) -> str:
    return "col:" + sheet + "/" + "/".join(header_path)


def _build_entity_index(wb) -> dict[str, dict]:
    """Map in-workbook entity name -> {kind, sheet, ref} for Excel Tables and
    named ranges, so PowerQuery's Excel.CurrentWorkbook references resolve to a
    real sheet/range. Tables take precedence over equally-named ranges."""
    idx: dict[str, dict] = {}
    for ws in wb.worksheets:
        try:
            for tname in ws.tables:
                tbl = ws.tables[tname]
                idx[tname] = {"kind": "table", "sheet": ws.title,
                              "ref": getattr(tbl, "ref", str(tbl))}
        except Exception:  # noqa: BLE001 - best-effort enrichment
            continue
    try:
        for name in wb.defined_names:
            dn = wb.defined_names[name]
            dests = list(dn.destinations)
            if dests and name not in idx:
                sheet, ref = dests[0]
                idx[name] = {"kind": "named_range", "sheet": sheet,
                             "ref": ref.replace("$", "")}
    except Exception:  # noqa: BLE001
        pass
    return idx


def analyze(xlsx_path: str) -> Model:
    model = Model(file=xlsx_path)

    # --- PowerQuery layer ---
    queries, pq_warnings = extract_powerquery(xlsx_path)
    for w in pq_warnings:
        model.warn(w)
    query_names = {q.name for q in queries}
    source_ids: set[str] = set()
    for q in queries:
        model.add_node(Node(id=f"query:{q.name}", type="query",
                            attrs={"m_code": q.m_code, "uris": q.uris}))
        uri_detail = ", ".join(q.uris)
        for src in q.sources:
            sid = f"source:{src}"
            if sid not in source_ids:
                model.add_node(Node(id=sid, type="source", attrs={"kind": "function"}))
                source_ids.add(sid)
            model.add_edge(Edge(source=sid, target=f"query:{q.name}",
                                edge_type="dataflow", granularity="pq",
                                detail=uri_detail))
        for ref in q.refs:
            model.add_edge(Edge(source=f"query:{ref}", target=f"query:{q.name}",
                                edge_type="dataflow", granularity="pq"))

        # Decompose the let-body into M steps (only worth showing for >=2).
        steps = decompose_steps(q.m_code)
        if len(steps) >= 2:
            for idx, st in enumerate(steps):
                model.add_node(Node(id=f"step:{q.name}/{st.name}", type="step",
                                    attrs={"query": q.name, "name": st.name,
                                           "index": idx, "expr": st.expr}))
            for st in steps:
                for ref in st.refs:
                    model.add_edge(Edge(source=f"step:{q.name}/{ref}",
                                        target=f"step:{q.name}/{st.name}",
                                        edge_type="dataflow", granularity="step"))

    # --- Workbook / sheets layer ---
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as exc:  # noqa: BLE001 - record, never crash
        model.warn(f"ワークブックを開けませんでした: {exc}")
        return model
    pq_targets = {name for name in wb.sheetnames if name in query_names}

    # Map sheet -> {col_index: LogicalColumn} for reference resolution.
    sheet_cols: dict[str, dict[int, list[str]]] = {}

    for ws in wb.worksheets:
        try:
            stype, detail = classify_sheet(ws, pq_targets)
            has_merged = bool(ws.merged_cells.ranges)
            model.add_node(Node(id=f"sheet:{ws.title}", type="sheet",
                                attrs={"sheet_type": stype, "has_merged": has_merged,
                                       "formula_ratio": round(detail["formula_ratio"], 3),
                                       "preview": sheet_preview(ws)}))
            cols = recover_headers(ws)
            sheet_cols[ws.title] = {c.col_index: c.header_path for c in cols}
            for c in cols:
                model.add_node(Node(
                    id=_col_id(ws.title, c.header_path), type="column",
                    attrs={"sheet": ws.title, "header_path": c.header_path,
                           "is_formula": c.is_formula, "confidence": c.confidence,
                           "preview": column_preview(ws, c.col_index, c.data_start)}))
        except Exception as exc:  # noqa: BLE001 - record, never crash
            model.warn(f"シート '{ws.title}' の解析に失敗: {exc}")

    # query -> sheet dataflow when a sheet name matches a query name
    for q in queries:
        if q.name in wb.sheetnames:
            model.add_edge(Edge(source=f"query:{q.name}", target=f"sheet:{q.name}",
                                edge_type="dataflow", granularity="pq"))

    # PowerQuery Excel.CurrentWorkbook references -> real table / named range,
    # resolved to the sheet they live on (slice ④ entity mapping).
    entity_idx = _build_entity_index(wb)
    sheet_ids = {f"sheet:{s.title}" for s in wb.worksheets}
    entity_ids: set[str] = set()
    for q in queries:
        for ent in q.wb_entities:
            eid = f"range:{ent}"
            info = entity_idx.get(ent)
            if eid not in entity_ids:
                attrs = {"name": ent, "kind": info["kind"] if info else "unresolved"}
                if info:
                    attrs.update(sheet=info["sheet"], ref=info["ref"])
                model.add_node(Node(id=eid, type="range", attrs=attrs))
                entity_ids.add(eid)
                if info and f"sheet:{info['sheet']}" in sheet_ids:
                    model.add_edge(Edge(source=f"sheet:{info['sheet']}", target=eid,
                                        edge_type="dataflow", granularity="pq",
                                        detail=info["ref"]))
            model.add_edge(Edge(source=eid, target=f"query:{q.name}",
                                edge_type="dataflow", granularity="pq"))

    # --- Cross-sheet references ---
    sheet_titles = {s.title for s in wb.worksheets}
    for ws in wb.worksheets:
        try:
            for ref in extract_references(ws):
                tgt_sheet = ref["target_sheet"]
                if tgt_sheet not in sheet_titles:
                    continue
                # try to resolve source cell's column -> logical column on this sheet
                src_col_idx = cell_to_column_index(ref["src_cell"])
                src_path = sheet_cols.get(ws.title, {}).get(src_col_idx)
                tgt_col_idx = cell_to_column_index(ref["target_ref"].split(":")[0])
                tgt_path = sheet_cols.get(tgt_sheet, {}).get(tgt_col_idx)

                if src_path and tgt_path:
                    model.add_edge(Edge(
                        source=_col_id(ws.title, src_path),
                        target=_col_id(tgt_sheet, tgt_path),
                        edge_type="reference", granularity="L2",
                        detail=ref["formula"]))
                else:
                    model.add_edge(Edge(
                        source=f"sheet:{ws.title}", target=f"sheet:{tgt_sheet}",
                        edge_type="reference", granularity="L1",
                        detail=ref["formula"]))
        except Exception as exc:  # noqa: BLE001 - record, never crash
            model.warn(f"シート '{ws.title}' の参照解析に失敗: {exc}")

    return model
