"""Worksheet-level analysis: type classification and (Task 4) header recovery."""
from __future__ import annotations

from dataclasses import dataclass
from openpyxl.utils import get_column_letter


@dataclass
class LogicalColumn:
    header_path: list[str]
    col_letter: str
    col_index: int
    is_formula: bool = False
    confidence: str = "high"
    sheet: str = ""


def _horizontal_merges(ws) -> dict[int, tuple[int, int, str]]:
    """Map each covered column index (1-based) on the merge's top row to
    (row, top_left_col, label) for horizontal merges spanning >1 column."""
    out: dict[int, tuple[int, int, str]] = {}
    for rng in ws.merged_cells.ranges:
        if rng.max_col > rng.min_col:  # horizontal span
            label = ws.cell(row=rng.min_row, column=rng.min_col).value
            label = "" if label is None else str(label)
            for col in range(rng.min_col, rng.max_col + 1):
                out[col] = (rng.min_row, rng.min_col, label)
    return out


def _column_has_formula(ws, col_index: int, start_row: int) -> bool:
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_index)
        if cell.data_type == "f":
            return True
    return False


def recover_headers(ws, max_header_rows: int = 3) -> list[LogicalColumn]:
    merges = _horizontal_merges(ws)
    has_merge = bool(merges)
    header_rows = 2 if has_merge else 1
    data_start = header_rows + 1

    cols: list[LogicalColumn] = []
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        parts: list[str] = []
        confidence = "low"
        if col in merges:
            _, _, parent = merges[col]
            if parent:
                parts.append(parent)
            child = ws.cell(row=header_rows, column=col).value
            if child is not None and str(child) != "":
                parts.append(str(child))
            confidence = "high"
        else:
            top = ws.cell(row=1, column=col).value
            if top is not None and str(top) != "":
                parts.append(str(top))
            if has_merge:
                child = ws.cell(row=header_rows, column=col).value
                if child is not None and str(child) != "" and (not parts or parts[-1] != str(child)):
                    parts.append(str(child))

        if not parts:
            continue
        is_formula = _column_has_formula(ws, col, data_start)
        cols.append(
            LogicalColumn(
                header_path=parts,
                col_letter=get_column_letter(col),
                col_index=col,
                is_formula=is_formula,
                confidence=confidence,
                sheet=ws.title,
            )
        )
    return cols


def classify_sheet(ws, pq_targets: set[str]) -> tuple[str, dict]:
    if ws.title in pq_targets:
        return "pq_output", {"formula_ratio": 0.0, "cell_count": 0}

    cell_count = 0
    formula_cells = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell_count += 1
            if cell.data_type == "f":
                formula_cells += 1

    ratio = (formula_cells / cell_count) if cell_count else 0.0
    detail = {"formula_ratio": ratio, "cell_count": cell_count}

    if formula_cells == 0:
        return "raw", detail
    if ratio > 0.5:
        return "formula", detail
    return "mixed", detail
