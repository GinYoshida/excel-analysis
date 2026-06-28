"""Capture small, capped previews of the actual cells behind a node.

Sheet and column nodes carry a `preview` payload so the HTML detail panel can
show real cell content (values + formulas + range address) instead of only the
recovered structure. Everything here is bounded so the embedded model JSON
cannot blow up on large sheets.
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter


def cell_display(cell) -> str:
    """Human-readable text for one cell: the formula (with '=') for formula
    cells, a clean integer for whole-number floats, else str(value)."""
    if cell.value is None:
        return ""
    if cell.data_type == "f":
        return str(cell.value)  # openpyxl keeps the leading '='
    v = cell.value
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def sheet_preview(ws, max_rows: int = 6, max_cols: int = 8) -> dict:
    n_rows = min(ws.max_row, max_rows)
    n_cols = min(ws.max_column, max_cols)
    grid: list[list[str]] = []
    for r in range(1, n_rows + 1):
        grid.append([cell_display(ws.cell(row=r, column=c))
                     for c in range(1, n_cols + 1)])
    return {
        "dims": {"rows": ws.max_row, "cols": ws.max_column},
        "truncated": ws.max_row > max_rows or ws.max_column > max_cols,
        "cells": grid,
    }


def column_preview(ws, col_index: int, data_start: int,
                   max_samples: int = 5) -> dict:
    """Sample non-empty values down one logical column plus a formula example
    and the column's data range address."""
    samples: list[str] = []
    formula_example: str | None = None
    for r in range(data_start, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_index)
        if cell.value is None:
            continue
        disp = cell_display(cell)
        if cell.data_type == "f" and formula_example is None:
            formula_example = disp
        if len(samples) < max_samples:
            samples.append(disp)
        if len(samples) >= max_samples and formula_example is not None:
            break
    letter = get_column_letter(col_index)
    return {
        "samples": samples,
        "formula_example": formula_example,
        "range": f"{letter}{data_start}:{letter}{ws.max_row}",
    }
